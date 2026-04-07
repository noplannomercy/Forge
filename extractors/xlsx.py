import io
from openpyxl import load_workbook

from models import ConvertResult, Quality


async def extract(file_bytes: bytes, file_name: str) -> ConvertResult:
    """XLSX → Markdown 변환 (시트별 표)"""
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    parts: list[str] = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_parts: list[str] = [f"## {sheet_name}"]

        rows_data = list(ws.iter_rows(values_only=True))
        if not rows_data:
            parts.append("\n\n".join(sheet_parts))
            continue

        # 헤더
        header = [str(c) if c is not None else "" for c in rows_data[0]]
        table_rows = ["| " + " | ".join(header) + " |"]
        table_rows.append("| " + " | ".join(["---"] * len(header)) + " |")

        # 데이터
        for row in rows_data[1:]:
            cells = [str(c) if c is not None else "" for c in row]
            table_rows.append("| " + " | ".join(cells) + " |")

        sheet_parts.append("\n".join(table_rows))
        parts.append("\n\n".join(sheet_parts))

    wb.close()

    full_text = "\n\n---\n\n".join(parts)
    total_chars = len(full_text.strip())
    num_sheets = len(wb.sheetnames)

    return ConvertResult(
        text=full_text,
        format="md",
        pages=num_sheets,
        file_name=file_name,
        source_format="xlsx",
        route="extract",
        quality=Quality(
            total_chars=total_chars,
            chars_per_page=total_chars / num_sheets if num_sheets > 0 else 0,
            total_pages=num_sheets,
            failed_pages=0,
            confidence="high",
        ),
    )
